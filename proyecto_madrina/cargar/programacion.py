from pathlib import Path

from openpyxl import load_workbook

import normalizar
from modelo import Clase, DIAS


COLUMNAS_ESPERADAS = [
    "N", "CATEDRATICO", "CODIGO", "ASIGNATURA", "CARRERA", "ALUMNOS",
    "MODALIDAD", "AULA", "SECCION", "HORAS PRESENCIALES", "HORAS ASINCRONICAS",
    "HORAS TOTALES", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM", "OBSERVACIONES",
]


def _es_fila_vacia(row) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


def _encontrar_fila_encabezado(ws, max_buscar: int = 15) -> int:
    """Localiza la fila con los nombres de columna. Espera 'CATEDR' en algún lado."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > max_buscar:
            break
        for cell in row:
            if cell and isinstance(cell, str) and "CATEDR" in cell.upper():
                return i
    return 7


CONSOLIDADAS_CONOCIDAS = {"PREGRADO-IA"}


def cargar_excel(
    path: str | Path,
    fuente: str = "PROGRAMACION",
    omitir_consolidadas: bool = True,
) -> list[Clase]:
    """Lee las hojas del Excel y devuelve una lista de Clase.

    Si `omitir_consolidadas` está activo, se saltan las hojas listadas en
    `CONSOLIDADAS_CONOCIDAS` cuando hay otras hojas que cargar. La hoja
    consolidada repite el contenido de las hojas por carrera y dispara
    260 falsos duplicados si se incluye."""
    wb = load_workbook(path, data_only=True, read_only=True)
    clases: list[Clase] = []

    hojas_a_leer = list(wb.sheetnames)
    if omitir_consolidadas and len(hojas_a_leer) > 1:
        hojas_a_leer = [h for h in hojas_a_leer if h not in CONSOLIDADAS_CONOCIDAS]

    for nombre_hoja in hojas_a_leer:
        ws = wb[nombre_hoja]
        fila_encabezado = _encontrar_fila_encabezado(ws)
        primera_data = fila_encabezado + 1
        vacias_consecutivas = 0

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i < primera_data:
                continue

            if _es_fila_vacia(row):
                vacias_consecutivas += 1
                if vacias_consecutivas >= 3:
                    break
                continue
            vacias_consecutivas = 0

            clase = _fila_a_clase(row, fuente=fuente, hoja=nombre_hoja, fila=i)
            if clase:
                clases.append(clase)

    return clases


def _fila_a_clase(row, fuente: str, hoja: str, fila: int) -> Clase | None:
    cols = list(row) + [None] * max(0, 20 - len(row))

    numero = cols[0] if isinstance(cols[0], (int, float)) else None
    catedratico, es_nuevo = normalizar.nombre_catedratico(cols[1])
    codigo, alternos = normalizar.codigos(cols[2])
    asignatura = normalizar.texto(cols[3]).upper()
    carrera = normalizar.texto(cols[4]).upper()

    if not catedratico and not codigo and not asignatura:
        return None

    alumnos = normalizar.alumnos(cols[5])
    modalidad = normalizar.texto(cols[6]).upper()
    aula = normalizar.aula(cols[7])
    seccion = normalizar.seccion(cols[8])
    h_pres = normalizar.horas(cols[9])
    h_asin = normalizar.horas(cols[10])
    h_tot = normalizar.horas(cols[11])

    bloques = []
    for idx, dia in enumerate(DIAS):
        bloques.extend(normalizar.bloques_dia(dia, cols[12 + idx]))

    observaciones = normalizar.texto(cols[19])

    return Clase(
        fuente=fuente,
        hoja=hoja,
        fila=fila,
        numero=int(numero) if numero is not None else None,
        catedratico=catedratico,
        catedratico_es_nuevo=es_nuevo,
        codigo=codigo,
        codigos_alternos=alternos,
        asignatura=asignatura,
        carrera=carrera,
        alumnos=alumnos,
        modalidad=modalidad,
        aula=aula,
        seccion=seccion,
        horas_presenciales=h_pres,
        horas_asincronicas=h_asin,
        horas_totales=h_tot,
        bloques=bloques,
        observaciones=observaciones,
    )
