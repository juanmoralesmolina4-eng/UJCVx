"""Exporta la lista de Clases normalizadas a un Excel con el mismo formato del
original, una hoja por carrera. Permite verificar a simple vista qué entendió
el parser y qué se va a enviar a Onlive/RRHH."""
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modelo import Clase, DIAS
from normalizar import minutos_a_hhmm


_AZUL = PatternFill("solid", fgColor="1F3A68")
_GRIS = PatternFill("solid", fgColor="F2F2F2")
_BLANCO_NEGRITA = Font(bold=True, color="FFFFFF")

ENCABEZADOS = [
    "Fila original",
    "Catedrático", "(NUEVO)",
    "Código", "Códigos alternos", "Asignatura", "Carrera",
    "Alumnos", "Modalidad", "Aula", "Sección",
    "Horas presenciales", "Horas asincrónicas", "Horas totales", "Horas calculadas",
    "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM",
    "Observaciones",
]

ANCHOS = [
    7,
    36, 9,
    14, 16, 38, 28,
    9, 14, 16, 10,
    11, 11, 11, 12,
    14, 14, 14, 14, 14, 14, 14,
    40,
]


def exportar(clases: list[Clase], ruta_salida: str | Path) -> None:
    por_hoja = defaultdict(list)
    for c in clases:
        por_hoja[c.hoja].append(c)

    wb = Workbook()
    wb.remove(wb.active)

    for hoja, lista in por_hoja.items():
        ws = wb.create_sheet(hoja[:31])
        _escribir_encabezado(ws)
        for i, c in enumerate(sorted(lista, key=lambda x: x.fila), start=2):
            _escribir_fila(ws, i, c)
        _aplicar_anchos(ws)
        ws.freeze_panes = "B2"

    wb.save(ruta_salida)


def _escribir_encabezado(ws) -> None:
    for col, texto in enumerate(ENCABEZADOS, 1):
        celda = ws.cell(row=1, column=col, value=texto)
        celda.font = _BLANCO_NEGRITA
        celda.fill = _AZUL
        celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _escribir_fila(ws, fila: int, c: Clase) -> None:
    bloques_por_dia = {d: [] for d in DIAS}
    for b in c.bloques:
        bloques_por_dia[b.dia].append(
            f"{minutos_a_hhmm(b.inicio)} - {minutos_a_hhmm(b.fin)}"
        )

    valores = [
        c.fila,
        c.catedratico, "Sí" if c.catedratico_es_nuevo else "",
        c.codigo, ", ".join(c.codigos_alternos), c.asignatura, c.carrera,
        c.alumnos, c.modalidad, c.aula, c.seccion,
        c.horas_presenciales, c.horas_asincronicas, c.horas_totales, c.horas_reales_semanales,
        *[" / ".join(bloques_por_dia[d]) for d in DIAS],
        c.observaciones,
    ]

    for col, val in enumerate(valores, 1):
        celda = ws.cell(row=fila, column=col, value=val)
        celda.alignment = Alignment(vertical="top", wrap_text=True)

    if fila % 2 == 0:
        for col in range(1, len(valores) + 1):
            ws.cell(row=fila, column=col).fill = _GRIS


def _aplicar_anchos(ws) -> None:
    for i, ancho in enumerate(ANCHOS, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
