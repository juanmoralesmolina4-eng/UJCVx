"""Exporta las métricas de eficiencia (aulas + docentes) a un Excel
con dos hojas. Pensado para que Martha pueda ordenar/filtrar y ver
de un vistazo dónde hay oportunidades de optimización."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from metricas.aulas import MetricasAula, HORAS_JORNADA_SEMANA
from metricas.docentes import MetricasDocente


_AZUL = PatternFill("solid", fgColor="1F3A68")
_GRIS = PatternFill("solid", fgColor="F2F2F2")
_ROJO = PatternFill("solid", fgColor="FCE4E4")
_VERDE = PatternFill("solid", fgColor="E4F7E4")
_BLANCO_NEGRITA = Font(bold=True, color="FFFFFF")


def exportar(
    metricas_aulas: list[MetricasAula],
    metricas_docentes: list[MetricasDocente],
    ruta_salida: str | Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_aulas(wb.create_sheet("Eficiencia de aulas"), metricas_aulas)
    _hoja_docentes(wb.create_sheet("Eficiencia docente"), metricas_docentes)
    wb.save(ruta_salida)


def _hoja_aulas(ws, metricas: list[MetricasAula]) -> None:
    encabezados = [
        "Aula", "Horas/semana", "% ocupación", "Clases", "Catedráticos",
        "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM",
    ]
    _escribir_encabezado(ws, encabezados)

    for i, m in enumerate(metricas, start=2):
        ws.cell(row=i, column=1, value=m.aula)
        ws.cell(row=i, column=2, value=round(m.horas_ocupadas, 1))
        ws.cell(row=i, column=3, value=round(m.porcentaje_ocupacion, 1))
        ws.cell(row=i, column=4, value=m.n_clases)
        ws.cell(row=i, column=5, value=len(m.catedraticos))
        for col, dia in enumerate(["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"], start=6):
            ws.cell(row=i, column=col, value=round(m.horas_por_dia.get(dia, 0), 1))

        if m.porcentaje_ocupacion < 25:
            for c in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=c).fill = _ROJO
        elif m.porcentaje_ocupacion > 80:
            for c in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=c).fill = _VERDE
        elif i % 2 == 0:
            for c in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=c).fill = _GRIS

    _anchos(ws, [22, 13, 12, 9, 13, 8, 8, 8, 8, 8, 8, 8])
    ws.freeze_panes = "B2"

    fila_ref = len(metricas) + 4
    ws.cell(row=fila_ref, column=1, value="Notas").font = Font(bold=True)
    ws.cell(row=fila_ref + 1, column=1,
            value=f"Jornada de referencia: 7:00–21:00 lun–sáb = {HORAS_JORNADA_SEMANA:.0f} h/semana.")
    ws.cell(row=fila_ref + 2, column=1, value="Rojo: ocupación < 25% (subutilizada). Verde: > 80% (saturada).")


def _hoja_docentes(ws, metricas: list[MetricasDocente]) -> None:
    encabezados = [
        "Catedrático", "Horas clase/sem", "Días/sem", "Clases",
        "Horas huecos/sem", "Eficiencia %", "Carreras", "Aulas",
    ]
    _escribir_encabezado(ws, encabezados)

    for i, m in enumerate(metricas, start=2):
        ws.cell(row=i, column=1, value=m.catedratico)
        ws.cell(row=i, column=2, value=round(m.horas_semanales, 1))
        ws.cell(row=i, column=3, value=m.n_dias)
        ws.cell(row=i, column=4, value=m.n_clases)
        ws.cell(row=i, column=5, value=round(m.horas_huecos, 1))
        ws.cell(row=i, column=6, value=round(m.ratio_eficiencia * 100, 1))
        ws.cell(row=i, column=7, value=", ".join(sorted(m.carreras)))
        ws.cell(row=i, column=8, value=", ".join(sorted(m.aulas)))

        if m.horas_huecos > 4 or m.ratio_eficiencia < 0.7:
            for c in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=c).fill = _ROJO
        elif i % 2 == 0:
            for c in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=c).fill = _GRIS

    _anchos(ws, [38, 14, 9, 8, 14, 12, 30, 25])
    ws.freeze_panes = "B2"

    fila_ref = len(metricas) + 4
    ws.cell(row=fila_ref, column=1, value="Notas").font = Font(bold=True)
    ws.cell(row=fila_ref + 1, column=1,
            value="Eficiencia = horas de clase ÷ horas en campus (clase + huecos). 100% = sin huecos.")
    ws.cell(row=fila_ref + 2, column=1,
            value="Rojo: docentes con huecos > 4h o eficiencia < 70%.")


def _escribir_encabezado(ws, encabezados: list[str]) -> None:
    for col, txt in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=txt)
        celda.font = _BLANCO_NEGRITA
        celda.fill = _AZUL
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def _anchos(ws, anchos: list[int]) -> None:
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
