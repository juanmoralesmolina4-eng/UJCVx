"""Genera el archivo de pago en el formato exacto que pide Recursos Humanos.

Basado en `PRIMER PAGO DOCENTE PREGRADO I PAC 2026 COMAYAGUA.xlsx`:
una hoja consolidada `PROGRAMACION GENERAL` + una hoja por carrera.

Fórmula del pago:
    horas_totales       = horas_por_semana × semanas_a_pagar
    valor_por_clase     = horas_totales × valor_por_hora_clase
    total_ingresos      = suma(valor_por_clase) por catedrático
    total_deducciones   = ihss + ujcv + embargo + ach
    total_a_pagar       = total_ingresos − total_deducciones

Las deducciones se dejan en 0 por defecto; RRHH las completa manualmente
o se calculan en una etapa posterior con datos del catedrático."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from modelo import Clase, DIAS


# Encabezado exacto del Excel de COMAYAGUA (34 columnas).
COLUMNAS = [
    "N°", "CATEDRÁTICO/A", "ASIGNATURA", "CARRERA", "CODIGO",
    "MODALIDAD", "AULA", "SECCIÓN", "ALUMNOS",
    "HORAS POR SEMANA", "HORAS ASINCRONICAS",
    "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM",
    "TIPO CONTRATACION", "OBSERVACIONES",
    "HORAS ASINCRONAS",
    "VALOR POR PERIODO COMPLETO",
    "SEMANAS A PAGAR",
    "HORAS POR SEMANA ",
    "HORAS TOTALES",
    "VALOR POR HORA CLASE", "VALOR POR CLASE",
    "TOTAL INGRESOS",
    "I.H.S.S", "UJCV", "EMBARGO", "ACH",
    "TOTAL DEDUCCIONES", "TOTAL A PAGAR",
]

_AZUL = PatternFill("solid", fgColor="1F3A68")
_AMARILLO = PatternFill("solid", fgColor="FFF3CD")
_BLANCO_NEGRITA = Font(bold=True, color="FFFFFF")
_NEGRITA = Font(bold=True)


@dataclass
class ParametrosPago:
    """Parámetros del cálculo de pago. Sobrescribir cuando vengan los
    catálogos de catedráticos/contratos."""
    semanas_a_pagar: int = config.SEMANAS_POR_PAGO_DEFAULT
    valor_por_hora_default: float = config.TARIFA_POR_HORA_DEFAULT_LPS
    tarifas_por_catedratico: dict[str, float] = field(default_factory=dict)
    tipo_contrato_por_catedratico: dict[str, str] = field(default_factory=dict)
    estipendio_por_catedratico: dict[str, float] = field(default_factory=dict)
    deducciones_por_catedratico: dict[str, dict[str, float]] = field(default_factory=dict)

    def tarifa_de(self, catedratico: str) -> float:
        return self.tarifas_por_catedratico.get(catedratico, self.valor_por_hora_default)

    def tipo_contrato_de(self, catedratico: str) -> str:
        return self.tipo_contrato_por_catedratico.get(catedratico, "")

    def estipendio_de(self, catedratico: str) -> float:
        return self.estipendio_por_catedratico.get(catedratico, 0.0)

    def deducciones_de(self, catedratico: str) -> dict[str, float]:
        return self.deducciones_por_catedratico.get(
            catedratico,
            {"ihss": 0.0, "ujcv": 0.0, "embargo": 0.0, "ach": 0.0},
        )


def exportar(
    clases: list[Clase],
    ruta_salida: str | Path,
    parametros: Optional[ParametrosPago] = None,
    titulo: str = "PRIMER PAGO DOCENTE",
    campus_nombre: str = "TEGUCIGALPA",
    periodo: str = "II PAC 2026",
) -> None:
    """Genera el archivo de pago en el formato de RRHH."""
    p = parametros or ParametrosPago()

    wb = Workbook()
    wb.remove(wb.active)

    ws_consol = wb.create_sheet("PROGRAMACION GENERAL")
    _escribir_titulo(ws_consol, titulo, campus_nombre, periodo)
    _escribir_encabezado(ws_consol, fila=7)
    _escribir_clases(ws_consol, clases, p, fila_inicio=8)
    _aplicar_anchos(ws_consol)

    por_carrera: dict[str, list[Clase]] = defaultdict(list)
    for c in clases:
        por_carrera[c.hoja].append(c)

    for hoja, lista in por_carrera.items():
        ws = wb.create_sheet(hoja[:31])
        _escribir_titulo(ws, titulo, campus_nombre, periodo, sub=hoja)
        _escribir_encabezado(ws, fila=7)
        _escribir_clases(ws, lista, p, fila_inicio=8)
        _aplicar_anchos(ws)

    wb.save(ruta_salida)


def _escribir_titulo(ws, titulo: str, campus: str, periodo: str, sub: str = "") -> None:
    ws.cell(row=2, column=2, value="UNIVERSIDAD JOSÉ CECILIO DEL VALLE").font = Font(bold=True, size=12)
    ws.cell(row=3, column=2, value=f"{titulo} {periodo}").font = _NEGRITA
    ws.cell(row=4, column=2, value=campus).font = _NEGRITA
    if sub:
        ws.cell(row=5, column=2, value=sub).font = Font(italic=True)


def _escribir_encabezado(ws, fila: int) -> None:
    for col, nombre in enumerate(COLUMNAS, 1):
        celda = ws.cell(row=fila, column=col, value=nombre)
        celda.font = _BLANCO_NEGRITA
        celda.fill = _AZUL
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[fila].height = 36
    ws.freeze_panes = ws.cell(row=fila + 1, column=3)


def _escribir_clases(
    ws,
    clases: list[Clase],
    p: ParametrosPago,
    fila_inicio: int,
) -> None:
    """Escribe las clases agrupadas por catedrático, con totales en la última
    fila de cada grupo (igual que el archivo original de COMAYAGUA)."""
    por_cated: dict[str, list[Clase]] = defaultdict(list)
    for c in clases:
        por_cated[c.catedratico].append(c)

    fila = fila_inicio
    n = 0
    for catedratico, grupo in sorted(por_cated.items()):
        n_grupo = len(grupo)
        deducciones = p.deducciones_de(catedratico)
        total_ingresos_grupo = sum(_valor_por_clase(c, p) for c in grupo)

        for idx, c in enumerate(grupo, start=1):
            n += 1
            es_ultima = idx == n_grupo
            _escribir_fila(
                ws, fila, n, c, p,
                total_ingresos=total_ingresos_grupo if es_ultima else None,
                deducciones=deducciones if es_ultima else None,
            )
            fila += 1


def _valor_por_clase(c: Clase, p: ParametrosPago) -> float:
    horas_semana = c.horas_presenciales or 0
    horas_totales = horas_semana * p.semanas_a_pagar
    return horas_totales * p.tarifa_de(c.catedratico)


def _escribir_fila(
    ws,
    fila: int,
    n: int,
    c: Clase,
    p: ParametrosPago,
    total_ingresos: Optional[float],
    deducciones: Optional[dict[str, float]],
) -> None:
    bloques_por_dia: dict[str, list[str]] = defaultdict(list)
    for b in c.bloques:
        h = f"{b.inicio // 60:02d}:{b.inicio % 60:02d}-{b.fin // 60:02d}:{b.fin % 60:02d}"
        bloques_por_dia[b.dia].append(h)

    horas_por_semana = c.horas_presenciales or 0
    horas_asincronicas = c.horas_asincronicas or 0
    semanas = p.semanas_a_pagar
    horas_totales = horas_por_semana * semanas
    tarifa = p.tarifa_de(c.catedratico)
    valor_por_clase = horas_totales * tarifa

    if deducciones is not None:
        ihss = deducciones["ihss"]
        ujcv = deducciones["ujcv"]
        embargo = deducciones["embargo"]
        ach = deducciones["ach"]
        total_deducciones = ihss + ujcv + embargo + ach
    else:
        ihss = ujcv = embargo = ach = total_deducciones = ""

    total_a_pagar = ""
    if total_ingresos is not None:
        total_a_pagar = total_ingresos - (total_deducciones or 0)

    valores = [
        n,
        c.catedratico,
        c.asignatura,
        c.carrera or c.hoja,
        c.codigo,
        c.modalidad,
        c.aula,
        c.seccion,
        c.alumnos if c.alumnos is not None else "",
        horas_por_semana,
        horas_asincronicas,
        *[" | ".join(bloques_por_dia.get(d, [])) for d in DIAS],
        p.tipo_contrato_de(c.catedratico),
        c.observaciones,
        horas_asincronicas,
        p.estipendio_de(c.catedratico),
        semanas,
        horas_por_semana,
        horas_totales,
        tarifa,
        valor_por_clase,
        total_ingresos if total_ingresos is not None else "",
        ihss, ujcv, embargo, ach,
        total_deducciones,
        total_a_pagar,
    ]

    for col, val in enumerate(valores, 1):
        celda = ws.cell(row=fila, column=col, value=val)
        celda.alignment = Alignment(vertical="top", wrap_text=True)

    if total_ingresos is not None:
        for col in (28, 33, 34):
            ws.cell(row=fila, column=col).font = _NEGRITA
            ws.cell(row=fila, column=col).fill = _AMARILLO


def _aplicar_anchos(ws) -> None:
    anchos = [
        5, 32, 32, 14, 12, 16, 14, 10, 9,
        10, 10,
        14, 14, 14, 14, 14, 14, 14,
        20, 30, 10,
        13, 9, 11, 11, 13, 12,
        13, 9, 9, 9, 9, 14, 14,
    ]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
