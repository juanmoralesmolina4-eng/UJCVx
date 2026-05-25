from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modelo import Problema


_AZUL = PatternFill("solid", fgColor="1F3A68")
_GRIS = PatternFill("solid", fgColor="F2F2F2")
_BLANCO_NEGRITA = Font(bold=True, color="FFFFFF")
_NEGRITA = Font(bold=True)

ETIQUETAS = {
    "consolidacion_inconsistente": "Consolidación inconsistente",
    "duplicado": "Duplicados",
    "solape_aula": "Solapes de aula",
    "solape_catedratico": "Solapes de catedrático",
    "horas_inconsistentes": "Horas inconsistentes",
    "horario_sospechoso": "Horarios sospechosos",
    "sobrecarga_docente": "Sobrecarga docente",
    "subutilizacion_docente": "Subutilización docente",
    "seccion_grande": "Secciones grandes",
}


def generar(problemas: list[Problema], ruta_salida: str | Path,
            total_clases: int) -> None:
    wb = Workbook()
    _hoja_resumen(wb.active, problemas, total_clases)

    por_tipo = defaultdict(list)
    for p in problemas:
        por_tipo[p.tipo].append(p)

    for tipo, lista in por_tipo.items():
        ws = wb.create_sheet(ETIQUETAS.get(tipo, tipo)[:31])
        _hoja_problemas(ws, lista)

    wb.save(ruta_salida)


def _hoja_resumen(ws, problemas: list[Problema], total_clases: int) -> None:
    ws.title = "Resumen"
    ws["A1"] = "Reporte de validación — Programación Académica II Período 2026"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Total de clases analizadas"
    ws["B3"] = total_clases
    ws["A3"].font = _NEGRITA

    ws["A4"] = "Total de problemas encontrados"
    ws["B4"] = len(problemas)
    ws["A4"].font = _NEGRITA

    conteo = Counter(p.tipo for p in problemas)
    ws["A6"] = "Tipo"
    ws["B6"] = "Cantidad"
    ws["A6"].font = _BLANCO_NEGRITA
    ws["B6"].font = _BLANCO_NEGRITA
    ws["A6"].fill = _AZUL
    ws["B6"].fill = _AZUL

    fila = 7
    for tipo, etiqueta in ETIQUETAS.items():
        ws.cell(row=fila, column=1, value=etiqueta)
        ws.cell(row=fila, column=2, value=conteo.get(tipo, 0))
        fila += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14


def _hoja_problemas(ws, problemas: list[Problema]) -> None:
    encabezados = ["#", "Severidad", "Descripción", "Hojas y filas afectadas"]
    for col, h in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=h)
        celda.font = _BLANCO_NEGRITA
        celda.fill = _AZUL
        celda.alignment = Alignment(horizontal="left", vertical="center")

    for i, p in enumerate(problemas, start=1):
        fila = i + 1
        ws.cell(row=fila, column=1, value=i)
        ws.cell(row=fila, column=2, value=p.severidad)
        ws.cell(row=fila, column=3, value=p.descripcion)
        refs_str = "; ".join(f"{h}!fila {f}" for _, h, f in p.referencias)
        ws.cell(row=fila, column=4, value=refs_str)

        if fila % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=fila, column=col).fill = _GRIS

        for col in range(1, 5):
            ws.cell(row=fila, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    anchos = [5, 11, 90, 50]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
