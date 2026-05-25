"""Verifica que la hoja consolidada coincida con la suma de las hojas por carrera."""
from openpyxl import load_workbook

from modelo import Problema
from cargar.programacion import CONSOLIDADAS_CONOCIDAS, _encontrar_fila_encabezado, _es_fila_vacia


def validar_archivo(path) -> list[Problema]:
    """Compara la cantidad de filas en cada hoja consolidada contra la suma
    de las hojas no consolidadas. Reporta si no cuadran."""
    wb = load_workbook(path, data_only=True, read_only=True)
    nombres = list(wb.sheetnames)

    consolidadas_presentes = [h for h in nombres if h in CONSOLIDADAS_CONOCIDAS]
    otras = [h for h in nombres if h not in CONSOLIDADAS_CONOCIDAS]

    if not consolidadas_presentes or not otras:
        return []

    suma_otras = sum(_contar_filas_datos(wb[h]) for h in otras)

    problemas: list[Problema] = []
    for hoja_consol in consolidadas_presentes:
        n_consol = _contar_filas_datos(wb[hoja_consol])
        if n_consol == suma_otras:
            continue
        problemas.append(Problema(
            tipo="consolidacion_inconsistente",
            severidad="alta",
            descripcion=(
                f"La hoja consolidada '{hoja_consol}' tiene {n_consol} registros, "
                f"pero la suma de las hojas por carrera ({', '.join(otras)}) suma {suma_otras}. "
                f"Diferencia: {n_consol - suma_otras:+d}."
            ),
            referencias=[("PROGRAMACION", hoja_consol, 0)],
            extra={
                "hoja_consolidada": hoja_consol,
                "registros_consolidada": n_consol,
                "registros_otras": suma_otras,
            },
        ))

    return problemas


def _contar_filas_datos(ws) -> int:
    fila_encabezado = _encontrar_fila_encabezado(ws)
    primera = fila_encabezado + 1
    vacias = 0
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < primera:
            continue
        if _es_fila_vacia(row):
            vacias += 1
            if vacias >= 3:
                break
            continue
        vacias = 0
        n += 1
    return n
