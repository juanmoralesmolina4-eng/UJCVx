import pandas as pd
from openpyxl import load_workbook

path = "PROGRAMACIÓN ACADÉMICA II PERÍODO 2026 PRESENCIAL.xlsx"

wb = load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    print(f"\n========== SHEET: {sheet_name} ==========")
    ws = wb[sheet_name]
    print(f"Dimensions: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
    # Print first 50 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 60:
            print(f"... ({ws.max_row - 60} more rows)")
            break
        # Trim None tail
        vals = list(row)
        while vals and vals[-1] is None:
            vals.pop()
        if vals:
            print(f"R{i}: {vals}")
