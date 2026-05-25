from openpyxl import load_workbook

path = "PROGRAMACIÓN ACADÉMICA II PERÍODO 2026 (1).xlsx"
wb = load_workbook(path, data_only=True, read_only=True)
print("SHEETS:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n========== SHEET: {sheet_name} ==========")
    print(f"max_row={ws.max_row}, max_col={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15:
            print(f"... (more rows)")
            break
        vals = list(row)
        while vals and vals[-1] is None:
            vals.pop()
        if vals:
            print(f"R{i}: {vals[:25]}")
